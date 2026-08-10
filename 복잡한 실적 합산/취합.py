# -*- coding: utf-8 -*-
"""
여러 팀의 서로 다른 엑셀 양식(표준 세로형 / 오프셋형 / 가로형 / 품목별 시트형)을 읽어
공통 스키마(기준월·팀·품목·계정·통화·현지금액·적용환율·원화금액)로 취합한다.

- 위치(행/열 번호)를 고정하지 않고 「계정」「Account」「품목」「통화」「구분」「원화환율」 같은
  표식 셀을 찾아 헤더를 잡는다.
- 팀 목록·컬럼 위치·시트 개수를 코드에 하드코딩하지 않는다 (다음 달 양식 변화 대응).
- 원본 파일은 읽기만 하고 수정하지 않는다.

사용법:
    python 취합.py                # 취합.py가 있는 폴더에서 *yyyy-mm 폴더 중 가장 최근 것을 자동 선택해 그 달만 취합
    python 취합.py 1차_2026-01     # 폴더를 직접 지정해 그 달만 취합
    python 취합.py --누적          # 산하 모든 *yyyy-mm 폴더를 처음부터 다시 읽어 누적 파일을 새로 만든다

월별 결과: 취합.py와 같은 폴더에 {yymm}_취합.xlsx 로 저장 (실행마다 그 달 파일만 새로 만든다. 다른 달 파일엔 영향 없음).
누적 결과: 실적_누적_{yymmdd}_{n}.xlsx 로 저장 (yymmdd는 실행한 날짜, n은 같은 날짜 안에서 실행할 때마다 1씩 증가).
          기존 누적 파일을 덮어쓰지 않고 매번 새 번호로 저장하므로, 이전 결과와 비교하려면 파일명으로 구분한다.
"""
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent
MONTHLY_SHEET = "취합"
CUMULATIVE_SHEET = "누적"
OUTPUT_HEADER = ["기준월", "팀", "품목", "계정", "통화", "현지금액", "적용환율", "원화금액", "정정여부"]
FOLDER_MONTH_RE = re.compile(r"(\d{4})-(\d{2})$")

REAL, SUBTOTAL, EXCLUDE = "REAL", "SUBTOTAL", "EXCLUDE"

# 계정명 매핑: 같은 뜻이지만 팀마다 다르게 적은 표현 -> (표준 명칭, 분류)
ALIASES = {
    "매출액": ("매출액", REAL),
    "매출": ("매출액", REAL),
    "revenue": ("매출액", REAL),
    "매출원가": ("매출원가", REAL),
    "cogs": ("매출원가", REAL),
    "판관비": ("판관비", REAL),
    "판매관리비": ("판관비", REAL),
    "판매비와관리비": ("판관비", REAL),
    "sg&a": ("판관비", REAL),
    "영업외손익": ("영업외손익", REAL),
    "영업외수지": ("영업외손익", REAL),
    "non-op income": ("영업외손익", REAL),
    "매출총이익": ("매출총이익", SUBTOTAL),
    "gross profit": ("매출총이익", SUBTOTAL),
    "영업이익": ("영업이익", SUBTOTAL),
    "operating income": ("영업이익", SUBTOTAL),
    "세전이익": ("세전이익", SUBTOTAL),
    "pre-tax income": ("세전이익", SUBTOTAL),
    "total": ("Total", EXCLUDE),
}

# 품목명이 아니라 "값 하나"라는 뜻으로만 쓰인 헤더 (품목별 시트형에서 시트명을 품목명으로 대신 씀)
GENERIC_VALUE_LABELS = {"amount", "금액", "값", "value"}

ACCOUNT_ANCHORS = {"계정", "account"}
ITEM_ANCHOR = "품목"
CURRENCY_CODE_RE = re.compile(r"\b(KRW|USD|IDR|EUR|JPY|CNY|GBP|SGD|HKD|THB|VND)\b")


def normalize(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def parse_amount(raw):
    """콤마·공백이 섞인 문자열, 괄호 표기 음수(회계식)를 숫자로 바꾼다."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace(",", "").replace(" ", "").replace(" ", "")
    try:
        val = float(s)
    except ValueError:
        return None
    if neg:
        val = -val
    return int(val) if val.is_integer() else val


def derive_month_from_name(name: str):
    m = re.search(r"(\d{4})-(\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return None


def derive_month_from_sheet(sheet_name: str):
    """시트 이름 자체가 월코드(YYMM, 예: 2601)인 경우 그 시트의 기준월로 삼는다."""
    if re.fullmatch(r"\d{4}", sheet_name.strip()):
        yy, mm = sheet_name[:2], sheet_name[2:]
        return f"20{yy}-{mm}"
    return None


def derive_team_name(file_stem: str) -> str:
    m = re.match(r"^(.*?)_(\d{3,4})$", file_stem)
    return m.group(1) if m else file_stem


def find_currency(ws):
    # 1) '통화' 라벨 셀 -> 옆/아래 값
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() == "통화":
                r, c = cell.row, cell.column
                for candidate in (ws.cell(row=r, column=c + 1).value, ws.cell(row=r + 1, column=c).value):
                    if candidate:
                        m = CURRENCY_CODE_RE.search(str(candidate).upper())
                        if m:
                            return m.group(1)
    # 2) "Currency: USD" / "통화: KRW" / "단위: KRW" 같은 문구
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            m = re.search(r"(?:Currency|통화|단위)\s*[:：]?\s*([A-Za-z]{3})", str(cell.value), re.IGNORECASE)
            if m and CURRENCY_CODE_RE.fullmatch(m.group(1).upper()):
                return m.group(1).upper()
    # 3) 어디든 통화 코드 하나가 단독으로 등장하는 경우 (예: "품목별 (USD)")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            m = CURRENCY_CODE_RE.search(str(cell.value).upper())
            if m:
                return m.group(1)
    return None


def find_fx_sheet(wb):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            labels = {str(c.value).strip() for c in row if c.value is not None}
            if {"통화", "구분", "원화환율"} <= labels:
                return ws
    return None


def parse_fx_rates(ws) -> dict:
    """환율표에서 통화별 '월평균' 환율만 뽑는다."""
    header_row, col = None, {}
    for row in ws.iter_rows():
        cols = {str(c.value).strip(): c.column for c in row if c.value is not None}
        if {"통화", "구분", "원화환율"} <= cols.keys():
            header_row = row[0].row
            col = {k: cols[k] for k in ("통화", "구분", "원화환율")}
            break
    rates = {}
    if header_row is None:
        return rates
    r = header_row + 1
    while r <= ws.max_row:
        currency = ws.cell(row=r, column=col["통화"]).value
        if currency is None or str(currency).strip() == "":
            break
        gubun = ws.cell(row=r, column=col["구분"]).value
        rate = ws.cell(row=r, column=col["원화환율"]).value
        if gubun and str(gubun).strip() == "월평균" and rate is not None:
            rates[str(currency).strip().upper()] = float(rate)
        r += 1
    return rates


def find_exact_cell(ws, targets: set):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and normalize(cell.value) in targets:
                return cell.row, cell.column
    return None


def resolve_rate(currency, month, fx_by_month, team, sheet_name, warnings):
    if currency == "KRW":
        return 1.0
    rate = fx_by_month.get(month, {}).get(currency)
    if rate is None:
        warnings.append(f"[{team}/{sheet_name}] 환율 정보 없음: 기준월 '{month}' 통화 '{currency}' - 이 시트는 건너뜁니다.")
    return rate


def make_record(month, team, item, account, currency, local_amt, rate, is_revision):
    krw = local_amt if currency == "KRW" else round(local_amt * rate)
    return {
        "기준월": month, "팀": team, "품목": item, "계정": account,
        "통화": currency, "현지금액": local_amt, "적용환율": rate, "원화금액": krw,
        "정정여부": "정정" if is_revision else "원본",
    }


def parse_sheet_vertical(ws, sheet_name, team, month, fx_by_month, warnings, is_revision):
    """표준 세로형 / 오프셋형 / 품목별 시트형 (모두 '계정' 또는 'Account' 표식이 앵커)."""
    anchor = find_exact_cell(ws, ACCOUNT_ANCHORS)
    if anchor is None:
        return None
    r0, c0 = anchor

    currency = find_currency(ws)
    if currency is None:
        warnings.append(f"[{team}/{sheet_name}] 통화를 찾지 못해 건너뜁니다.")
        return []
    rate = resolve_rate(currency, month, fx_by_month, team, sheet_name, warnings)
    if rate is None:
        return []

    # 품목 헤더 행 찾기: 오른쪽에 뭔가 있고, 그 바로 다음 행의 계정열이 실제 계정명이면 그 행이 헤더다.
    item_header_row, data_start_row = None, None
    for hr in range(r0, r0 + 5):
        has_right = any(
            ws.cell(row=hr, column=c).value not in (None, "")
            for c in range(c0 + 1, ws.max_column + 1)
        )
        if not has_right:
            continue
        acct_val = ws.cell(row=hr + 1, column=c0).value
        if acct_val is not None and normalize(acct_val) in ALIASES:
            item_header_row, data_start_row = hr, hr + 1
            break
    if item_header_row is None:
        warnings.append(f"[{team}/{sheet_name}] 품목 헤더를 찾지 못했습니다 (데이터 없음으로 처리).")
        return []

    item_cols = [
        (c, str(ws.cell(row=item_header_row, column=c).value).strip())
        for c in range(c0 + 1, ws.max_column + 1)
        if ws.cell(row=item_header_row, column=c).value not in (None, "")
    ]
    if len(item_cols) == 1 and item_cols[0][1].lower() in GENERIC_VALUE_LABELS:
        item_cols = [(item_cols[0][0], sheet_name.strip())]

    records = []
    row = data_start_row
    while row <= ws.max_row:
        acct_raw = ws.cell(row=row, column=c0).value
        if acct_raw is None or str(acct_raw).strip() == "":
            break
        info = ALIASES.get(normalize(acct_raw))
        if info is None:
            warnings.append(f"[{team}/{sheet_name}] 매핑에 없는 계정명: '{acct_raw}' (행 {row}) - 건너뜁니다.")
            row += 1
            continue
        canon, category = info
        if category == REAL:
            for col, item_name in item_cols:
                amt = parse_amount(ws.cell(row=row, column=col).value)
                if amt is not None:
                    records.append(make_record(month, team, item_name, canon, currency, amt, rate, is_revision))
        row += 1
    return records


def parse_sheet_horizontal(ws, sheet_name, team, month, fx_by_month, warnings, is_revision):
    """가로형: 행 = 품목, 열 = 계정 ('품목' 표식이 앵커)."""
    anchor = find_exact_cell(ws, {ITEM_ANCHOR})
    if anchor is None:
        return None
    r0, c0 = anchor

    currency = find_currency(ws)
    if currency is None:
        warnings.append(f"[{team}/{sheet_name}] 통화를 찾지 못해 건너뜁니다.")
        return []
    rate = resolve_rate(currency, month, fx_by_month, team, sheet_name, warnings)
    if rate is None:
        return []

    acct_cols = []
    for c in range(c0 + 1, ws.max_column + 1):
        header = ws.cell(row=r0, column=c).value
        if header is None or str(header).strip() == "":
            continue
        info = ALIASES.get(normalize(header))
        if info is None:
            warnings.append(f"[{team}/{sheet_name}] 매핑에 없는 계정명: '{header}' (열 {c}) - 건너뜁니다.")
            continue
        canon, category = info
        if category == REAL:
            acct_cols.append((c, canon))

    records = []
    row = r0 + 1
    while row <= ws.max_row:
        item = ws.cell(row=row, column=c0).value
        if item is None or str(item).strip() == "":
            break
        item = str(item).strip()
        for col, canon in acct_cols:
            amt = parse_amount(ws.cell(row=row, column=col).value)
            if amt is not None:
                records.append(make_record(month, team, item, canon, currency, amt, rate, is_revision))
        row += 1
    return records


def parse_sheet(ws, sheet_name, team, month, fx_by_month, warnings, is_revision):
    for parser in (parse_sheet_vertical, parse_sheet_horizontal):
        result = parser(ws, sheet_name, team, month, fx_by_month, warnings, is_revision)
        if result is not None:
            return result
    warnings.append(f"[{team}/{sheet_name}] 알려진 양식(세로형/오프셋형/가로형/품목별 시트형) 중 어디에도 맞지 않아 건너뜁니다.")
    return []


def load_all_fx_rates():
    """모든 *yyyy-mm 폴더의 환율표를 각자의 폴더 기준월에 묶어 {기준월: {통화: 환율}} 로 만든다.
    (팀 파일 안에 다른 달 시트가 섞여 있어도 그 달 고유의 환율을 찾아 쓰기 위함)"""
    fx_by_month = {}
    for yyyy, mm, folder in find_month_folders(BASE_DIR):
        month = f"{yyyy}-{mm}"
        for f in sorted(folder.glob("*.xlsx")):
            wb = openpyxl.load_workbook(f, data_only=True)
            fx_sheet = find_fx_sheet(wb)
            if fx_sheet is not None:
                fx_by_month.setdefault(month, {}).update(parse_fx_rates(fx_sheet))
    return fx_by_month


def process_folder(folder: Path, fx_by_month: dict):
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        print(f"{folder}에서 xlsx 파일을 찾을 수 없습니다.", file=sys.stderr)
        sys.exit(1)

    folder_month = derive_month_from_name(folder.name)

    team_files = []
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        if find_fx_sheet(wb) is None:
            team_files.append((f, wb))

    records, warnings, read_log = [], [], []
    team_row_counts = defaultdict(int)
    for f, wb in team_files:
        team = derive_team_name(f.stem)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_own_month = derive_month_from_sheet(sheet_name)
            month = sheet_own_month or folder_month
            if month is None:
                warnings.append(f"[{team}/{sheet_name}] 기준월을 판별하지 못해 건너뜁니다.")
                continue
            # 시트명이 스스로 밝힌 기준월이 이 폴더 자신의 월과 다르면, 지난달 실적이 이후 파일에
            # 다시 실려 온 것 -> 정정본으로 표시한다 (원본과 구분만 하고 지우지는 않는다).
            is_revision = sheet_own_month is not None and sheet_own_month != folder_month
            parsed = parse_sheet(ws, sheet_name, team, month, fx_by_month, warnings, is_revision)
            records.extend(parsed)
            team_row_counts[team] += len(parsed)
            read_log.append((f.name, sheet_name, len(parsed)))

    return records, warnings, read_log, team_row_counts


def write_output(records, output_file: Path, sheet_name: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(OUTPUT_HEADER)
    for rec in records:
        ws.append([rec[col] for col in OUTPUT_HEADER])
    wb.save(output_file)


def next_cumulative_output_path() -> Path:
    """실적_누적_yymmdd_n.xlsx : 실행 날짜 + 그날 안에서의 실행 순번. 기존 파일을 덮어쓰지 않는다."""
    today_tag = date.today().strftime("%y%m%d")
    pattern = re.compile(rf"^실적_누적_{today_tag}_(\d+)\.xlsx$")
    used = [int(m.group(1)) for p in BASE_DIR.glob(f"실적_누적_{today_tag}_*.xlsx") if (m := pattern.match(p.name))]
    n = max(used) + 1 if used else 1
    return BASE_DIR / f"실적_누적_{today_tag}_{n}.xlsx"


def find_month_folders(base: Path):
    """base 바로 아래에서 이름이 ...yyyy-mm 형태로 끝나는 폴더를 찾는다."""
    found = []
    for p in base.iterdir():
        if p.is_dir():
            m = FOLDER_MONTH_RE.search(p.name)
            if m:
                found.append((m.group(1), m.group(2), p))
    return found


def pick_latest_month_folder(base: Path):
    found = find_month_folders(base)
    if not found:
        return None
    found.sort(key=lambda t: (t[0], t[1]))
    return found[-1][2]


def resolve_target_folder(folder_arg):
    if folder_arg:
        folder = Path(folder_arg)
        if not folder.is_absolute():
            folder = BASE_DIR / folder
        if not folder.is_dir():
            print(f"폴더를 찾을 수 없습니다: {folder}", file=sys.stderr)
            sys.exit(1)
        return folder

    folder = pick_latest_month_folder(BASE_DIR)
    if folder is None:
        print(f"{BASE_DIR}에서 '...yyyy-mm' 형태의 폴더를 찾을 수 없습니다. 폴더명을 인자로 지정하세요.", file=sys.stderr)
        sys.exit(1)
    print(f"[자동 선택] {folder.name}")
    return folder


def print_warnings(warnings):
    if warnings:
        print()
        print("=== 경고 ===")
        for w in warnings:
            print(f"  ! {w}")


def run_single(folder_arg):
    folder = resolve_target_folder(folder_arg)

    folder_month = derive_month_from_name(folder.name)
    if folder_month is None:
        print(f"폴더명에서 기준월(yyyy-mm)을 찾을 수 없습니다: {folder.name}", file=sys.stderr)
        sys.exit(1)
    yymm = folder_month[2:4] + folder_month[5:7]
    output_file = BASE_DIR / f"{yymm}_취합.xlsx"

    fx_by_month = load_all_fx_rates()
    records, warnings, read_log, team_row_counts = process_folder(folder, fx_by_month)

    # 이 폴더 안에 다른 달로 인식된 시트(예: 다음 달 파일 안에 섞여 온 지난달 시트)가 있으면
    # 이번 달 산출물에는 담지 않는다 (누적 모드에서는 그대로 담긴다).
    own_month = [r for r in records if r["기준월"] == folder_month]
    other_month = [r for r in records if r["기준월"] != folder_month]
    if other_month:
        by_month = defaultdict(int)
        for r in other_month:
            by_month[r["기준월"]] += 1
        for month, n in sorted(by_month.items()):
            warnings.append(f"[정보] 이 폴더에 기준월 '{month}' 데이터가 {n}건 섞여 있어 이번 달 파일에서는 제외했습니다 (누적 모드에서는 포함됩니다).")
        team_row_counts = defaultdict(int)
        for r in own_month:
            team_row_counts[r["팀"]] += 1

    write_output(own_month, output_file, MONTHLY_SHEET)

    print(f"[대상 폴더] {folder.name}")
    print()
    print("=== 읽은 파일 목록 ===")
    for filename, sheet_name, n in read_log:
        print(f"  - {filename} [{sheet_name}] : {n}행")
    print()
    print("=== 팀별 행수 ===")
    for team, n in sorted(team_row_counts.items()):
        print(f"  - {team}: {n}행")
    print()
    print(f"총 행수: {len(own_month)}행")
    print(f"[출력] {output_file}")
    print_warnings(warnings)


def run_cumulative():
    folders = [p for _, _, p in sorted(find_month_folders(BASE_DIR))]
    if not folders:
        print(f"{BASE_DIR}에서 '...yyyy-mm' 형태의 폴더를 찾을 수 없습니다.", file=sys.stderr)
        sys.exit(1)

    fx_by_month = load_all_fx_rates()
    all_records, all_warnings, read_log = [], [], []
    team_row_counts = defaultdict(int)
    for folder in folders:
        records, warnings, folder_read_log, folder_team_counts = process_folder(folder, fx_by_month)
        all_records.extend(records)
        all_warnings.extend(warnings)
        read_log.extend((folder.name, filename, sheet_name, n) for filename, sheet_name, n in folder_read_log)
        for team, n in folder_team_counts.items():
            team_row_counts[team] += n

    month_row_counts = defaultdict(int)
    for rec in all_records:
        month_row_counts[rec["기준월"]] += 1

    output_file = next_cumulative_output_path()
    write_output(all_records, output_file, CUMULATIVE_SHEET)

    print("[누적 모드] 아래 폴더를 모두 처음부터 다시 읽어 집계합니다:")
    for folder in folders:
        print(f"  - {folder.name}")
    print()
    print("=== 읽은 파일 목록 ===")
    for folder_name, filename, sheet_name, n in read_log:
        print(f"  - [{folder_name}] {filename} [{sheet_name}] : {n}행")
    print()
    print("=== 월별 행수 ===")
    for month, n in sorted(month_row_counts.items()):
        print(f"  - {month}: {n}행")
    print()
    print("=== 팀별 행수 ===")
    for team, n in sorted(team_row_counts.items()):
        print(f"  - {team}: {n}행")
    print()
    print(f"총 행수: {len(all_records)}행")
    print(f"[출력] {output_file}")
    print_warnings(all_warnings)


def main():
    args = sys.argv[1:]
    if "--누적" in args:
        run_cumulative()
    else:
        run_single(args[0] if args else None)


if __name__ == "__main__":
    main()
